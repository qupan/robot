#coding=utf-8
import xml.etree.ElementTree as ET
import xml.dom.minidom
import os,time,shutil,xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
try:
	import sys
	reload(sys)
	sys.setdefaultencoding('utf-8')
except Exception as e:
	print('run for python3')


class Excel_Report():

	def __init__(self,project_name):
		#得到报告的地址
		self.project_name=project_name
		self.report_time=time.strftime('%Y.%m.%d.%H.%M.%S')
		self.x=os.getcwd().split('word')[0]#得到当前的地址，以word分割
		self.output=self.x+'report\\output.xml'#得到output.xml文件地址
		self.log=self.x+'report\\log.html'#得到log.html文件地址
		self.report=self.x+'report\\report.html'#得到report.html文件地址
		self.html_addr='file:///'+self.x+'report\\report.html'#得到html文件地址

	def create_new_report(self):
		self.excel_addr = self.x + 'report\\excel_report'#得到excel报告的存放地址
		self.log_addr = self.x + 'report\\log'#得到log文本的存放地址

		self.sys_report_new_addr=self.x+'report\\sys_report'#得到新地址
		self.sys_report_addr=self.x+'report\\sys_report\\report'#得到新的report地址
		#print(self.sys_report_addr)
		self.sys_report_log=self.x+'report\\sys_report\\log'#得到新的log地址
		self.sys_report_xml=self.x+'report\\sys_report\\xml'#得到新的xml地址

		if os.path.exists(self.excel_addr) == False:
			os.mkdir(self.excel_addr)
		if os.path.exists(self.log_addr)==False:
			os.mkdir(self.log_addr)
		if os.path.exists(self.sys_report_new_addr)==False:
			os.mkdir(self.sys_report_new_addr)
		if os.path.exists(self.sys_report_addr)==False:
			os.mkdir(self.sys_report_addr)
		if os.path.exists(self.sys_report_log)==False:
			os.mkdir(self.sys_report_log)
		if os.path.exists(self.sys_report_xml)==False:
			os.mkdir(self.sys_report_xml)

		#新建报告存放地址保存每次生成的报告
		x=self.sys_report_addr+'\\'+self.report_time+'-Report.html'
		y=self.sys_report_log+'\\'+self.report_time+'-log.html'
		z=self.sys_report_xml+'\\'+self.report_time+'-output.xml'
		os.system('copy %s %s'%(self.report,x))#复制新的report.html
		os.system('copy %s %s'%(self.log,y))#复制新的log.html
		os.system('copy %s %s'%(self.output,z))#复制新的output.xml

		self.screenshot=os.path.join(self.x,'report\\screenshot')#截图文件夹
		self.screenshot_child=os.path.join(self.screenshot,self.report_time)#得到截图子文件夹
		report=os.path.join(self.x,'report')#得到report文件夹地址
		#创建截图的文件夹
		if os.path.exists(self.screenshot)==False:
			os.mkdir(self.screenshot)
		if os.path.exists(self.screenshot_child)==False:
			os.mkdir(self.screenshot_child)
		#移动错误截图分开存放，方便管理
		file_name=os.listdir(report)
		for i in file_name:
			if '.png' in i:
				name=os.path.join(report,i)
				shutil.move(name,self.screenshot_child)
		


	def create_excel(self):

		#创建新excel，加入格式字体，填充，颜色，预置名称等

		letter=['A','C','D','E','F','G','H','I']
		
		self.excel_name=self.excel_addr+'\\'+self.report_time+'-Report.xlsx'
		wb=Workbook()
		ws=wb['Sheet']
		ws.title=u'自动化测试报告'
		ws=wb.create_sheet(u'用例执行情况')

		#编辑第一个工作表格信息
		ws=wb[u'自动化测试报告']
		white = Font(size=12, bold=False, name=u'等线', color="FFFFFF")#白色字体
		black = Font(size=12, bold=False, name=u'等线', color="000000")#黑色字体
		#合并单元格
		ws.merge_cells('A1:I1')
		ws.merge_cells('A2:I2')
		ws.merge_cells('A5:I5')
		ws.merge_cells('A6:D6')
		ws.merge_cells('F6:I6')
		ws.merge_cells('A4:B4')
		for i in range(7,51):
			ws.merge_cells('B%d:C%d'%(i,i))
		#更改行高和列宽
		ws.column_dimensions['A'].width=5
		ws.column_dimensions['B'].width=15
		ws.column_dimensions['C'].width=18
		ws.column_dimensions['D'].width=25
		ws.column_dimensions['E'].width=15
		ws.column_dimensions['F'].width=20
		ws.column_dimensions['G'].width=15
		ws.column_dimensions['H'].width=25
		ws.column_dimensions['I'].width=25
		for i in range(1,51):
			ws.row_dimensions[i].height=15
		#添加表格线条，
		thin = Side(border_style="thin", color="000000")
		border = Border(left=thin, right=thin, top=thin, bottom=thin)
		for i in ['A','B','C','D','E','F','G','H','I']:
			for j in range(1,51):
				ws[i+str(j)].border=border
				alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)#所有表格居中
				ws[i+str(j)].alignment = alignment
				ws[i+str(j)].font=black
		#添加表格颜色
		ws['A1'].fill=PatternFill(patternType='solid',start_color='333399')
		ws['A2'].fill=PatternFill(patternType='solid',start_color='99CC00')
		for i in ['A','B','C','D','E','F','G','H','I']:
			ws[i+str(3)].fill=PatternFill(patternType='solid',start_color='0066CC')
			ws[i+str(3)].font=white
		ws['C6'].fill=PatternFill(patternType='solid',start_color='0066CC')
		for i in ['A','B','D']:
			ws[i+str(7)].fill=PatternFill(patternType='solid',start_color='0066CC')
		ws['A6'].fill=PatternFill(patternType='solid',start_color='99CC00')
		ws['E6'].fill=PatternFill(patternType='solid',start_color='FF6600')
		ws['E7'].fill=PatternFill(patternType='solid',start_color='FF9900')
		ws['F6'].fill=PatternFill(patternType='solid',start_color='FFCC99')
		ws['F7'].fill=PatternFill(patternType='solid',start_color='FFCC00')
		for i in ['F','G','H','I']:
			ws[i+str(7)].fill=PatternFill(patternType='solid',start_color='0066CC')
		#写入字符
		ws['A7'].font=ws['B7'].font=ws['D7'].font=ws['A1'].font=white
		ws['A1']=u'自动化测试报告'
		ws['F6'].font=ws['E6'].font=ws['E7'].font=ws['A6'].font=ws['A2'].font=black
		ws['F7'].font=ws['G7'].font=ws['H7'].font=ws['I7'].font=white
		ws['A2']=u'自动化测试用例总计'
		ws['A6']=u'测试模块列表'
		ws['E6']=u'是否执行'
		ws['E7']=u'执行标识符'
		ws['F6']=u'执行情况'
		ws['A7']=u'NO.'
		ws['B7']=u'模块'
		ws['D7']=u'自动化测试用例数量'
		ws['F7']=u'执行时间'
		ws['G7']=u'总计执行'
		ws['H7']=u'通过数目'
		ws['I7']=u'失败数目'
		title1=[u'自动化测试项目',u'自动化测试用例数量',u'自动化测试总计执行','Pass',
				'Fail',u'测试用时',u'开始时间',u'结束时间']
		ws.merge_cells('A3:B3')
		for i in range(len(letter)):
			ws[letter[i]+str(3)]=title1[i]

		#编辑第二个表格信息
		ws1=wb[u'用例执行情况']
		#更改单元格宽度
		for i in range(1,1000):
			ws1.row_dimensions[i].height=40#更改行高为40
		for i in ['A','B','C','D','E','F','G','H','I','J']:
			for j in range(1,1000):
				ws1[i+str(j)].border=border
				ws1[i+str(j)].font=black
				alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)#所有表格居中
				ws1[i+str(j)].alignment = alignment

		ws1.column_dimensions['A'].width=25
		ws1.column_dimensions['B'].width=40
		ws1.column_dimensions['C'].width=13
		ws1.column_dimensions['D'].width=35
		ws1.column_dimensions['E'].width=25
		ws1.column_dimensions['F'].width=40
		ws1.column_dimensions['G'].width=15
		ws1.column_dimensions['H'].width=100
		ws1.column_dimensions['I'].width=25
		ws1.column_dimensions['J'].width=25
		#合并单元格A1。添加颜色
		ws1.merge_cells('A1:J1')
		ws1['A1'].fill=PatternFill(patternType='solid',start_color='0066CC')
		ws1['A1'].alignment = alignment
		ws1['A1'].font=white
		ws1['A1']=u'自动化测试用例执行情况'
		#添加A2-G2格式和字体
		x1=['Suite Name','Case Name','Tags',u'输入数据',u'实际结果',u'预期结果','Status','Message','Start-Time','End-Time']
		x2=['A','B','C','D','E','F','G','H','I','J']
		for i in ['A','B','C','D','E','F','G','H','I','J']:
			ws1[i+str(2)].fill=PatternFill(patternType='solid',start_color='C0C0C0')#设置字体
		ws1['A2']=x1[0]
		ws1['B2']=x1[1]
		ws1['C2']=x1[2]
		ws1['D2']=x1[3]
		ws1['E2']=x1[4]
		ws1['F2']=x1[5]
		ws1['G2']=x1[6]
		ws1['H2']=x1[7]
		ws1['I2']=x1[8]
		ws1['J2']=x1[9]

		#wb.save(addr)#保存excel文件
		return wb#excel对象

	def total_time(self):
		#得到运行总时间
		chrome_options=webdriver.ChromeOptions()
		chrome_options.add_argument('--no-sandbox')#解决DevToolsActivePort文件不存在的报错
		chrome_options.add_argument('--headless')#浏览器不提供可视化页面. linux下如果系统不支持可视化不加这条会启动失败
		chrome_options.add_argument('--disable-gpu')#谷歌文档提到需要加上这个属性来规避bug
		chrome_options.add_argument('--window-size=1920,1080')#指定浏览器分辨率
		chrome_options.add_argument('blink-settings=imagesEnabled=false') #不加载图片, 提升速度
		chrome_options.add_argument('--hide-scrollbars') #隐藏滚动条, 应对一些特殊页面
		#chrome_options.binary_location = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe" #手动指定使用的浏览器位置
		driver=webdriver.Chrome(chrome_options=chrome_options)
		driver.get(self.html_addr)
		time.sleep(3)
		try:
			total_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(4) > td').text#得到总的时间
			start_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(2) > td').text#得到总的开始时间
			end_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(3) > td').text#得到总的结束时间
			tr=driver.find_elements_by_css_selector('#suite-stats > tbody>tr')
			message=[]
			s='::'
			for i in range(1,len(tr)+1):
				name=driver.find_element_by_css_selector('#suite-stats > tbody>tr:nth-child(%d)'%i)
				suit_name=name.find_element_by_css_selector('td:nth-child(1)').text#得到名字
				total_num=name.find_element_by_css_selector('td:nth-child(2)').text#得到用例集合的总数
				pass_num = name.find_element_by_css_selector('td:nth-child(3)').text#得到成功的数量
				file_num = name.find_element_by_css_selector('td:nth-child(4)').text#得到失败的数量
				run_time = name.find_element_by_css_selector('td:nth-child(5)').text#得到运行时间
				total_str=suit_name+s+total_num+s+pass_num+s+file_num+s+run_time
				message.append(total_str)
			project =total_time+s+start_time+s+end_time
			message.append(project)
			#print(message)
			driver.close()
			os.system('taskkill /f /im geckodriver.exe')
			os.system('taskkill /f /im chromedriver.exe')
			os.system('taskkill /f /im IEDriverServer.exe')
		except Exception as e:
			driver.close()
			os.system('taskkill /f /im geckodriver.exe')
			os.system('taskkill /f /im chromedriver.exe')
			os.system('taskkill /f /im IEDriverServer.exe')
		return message

	def read_xml(self):
		dom=xml.dom.minidom.parse(self.output)
		root=dom.documentElement
		suite=root.getElementsByTagName("suite")
		test=root.getElementsByTagName('test')
		#print(len(test))
		s='::'
		suit_message = []
		self.log_name=self.log_addr+'\\'+self.report_time+'-log.txt'
		file=open(self.log_name,'w')#创建新的log日志文件
		for i in suite:
			if '.txt' in i.getAttribute('source') or '.robot' in i.getAttribute('source') or '.tsv' in i.getAttribute('source') or'.html' in i.getAttribute('source'):
				suite_name=i.getAttribute('name')#得到集合的名字
				suit_message.append(suite_name)
				suit_status=i.getElementsByTagName('status')
				test_case=i.getElementsByTagName('test')#得到集合下面所有的用例放入列表中
				#print(len(test_case))#打印集合下面用例的数量
				for j in test_case:
					test_name=j.getAttribute('name')#得到用例名字
					#print(test_name)
					x=j.getElementsByTagName('status')#得到用例下面status标签
					case_status=x[-1].getAttribute('status')#得到用例的运行结果
					#print(case_status)
					endtime=x[-1].getAttribute('endtime')#得到用例结束时间
					starttime=x[-1].getAttribute('starttime')#得到用例开始开始时间
					file.write(suite_name+s)#写入集合名字
					file.write(test_name+s)#写入用例名字

					tags=j.getElementsByTagName('tag')#得到所有tag
					if len(tags)>0:#判断是否有tags标签
						tag=[]
						for k in tags:
							tag.append(k.firstChild.data)#把所有标签加入tag空列表中
						tag=','.join(tag)#将所有的标签链接在一起
						file.write(tag+s)#写入tag名称
					elif len(tags)<=0:
						file.write(u'空'+s)

					#输入数据一栏
					input_message=j.getElementsByTagName('kw')
					input_data=[]
					for i in input_message:
						if i.getAttribute('name')=='Input Text':
							msg=i.getElementsByTagName('msg')
							text=msg[0].firstChild.data
							file.write(text.split(' ')[2][1:-1]+',')
						elif i.getAttribute('name')=='Convert Dict':
							msg=i.getElementsByTagName('msg')
							text=msg[0].firstChild.data
							file.write(text[12:38]+',')
						else:
							input_data.append(u'空')
					if len(input_data) == len(input_message):
						file.write(u'未得到结果')
					file.write(s)

					#得到实际结果和预期结果
					kw=j.getElementsByTagName('kw')
					Log=[]
					for i in kw:
						if i.getAttribute('name')=='Log':
							#第一个是得到的信息，是实际结果
							x=i.getElementsByTagName('msg')
							y=x[0].firstChild.data
							file.write(y+s)#写入实际结果
						else:
							Log.append(u'未得到结果')
					if len(Log)==len(kw):#判断是否得到的实际结果，未得到就是脚本问题
						file.write(u'脚本运行失败，未得到实际结果' + s)#写入			
					file.write('message'+s)#写入实际结果，统一为message，后续读取每个模块的预期结果统一重写

					file.write(case_status + s)#写入用例的状态

					if case_status=='FAIL':
						Fail_message=x[-1].firstChild.data#如果是失败的用例，得到失败的信息
						print(Fail_message)
						file.write(Fail_message.replace('\n',',')+s)#失败信息取消空格
					elif case_status=='PASS':
						file.write(u'无' + s)#如果是通过的用例写入空格

					file.write(starttime + s + endtime + s)#写入开始时间和结束时间
					file.write(str(len(test_case)) + s)
					file.write('\n')
		file.close()
		return suit_message

	def read_all_excel(self):
		#读取出自动化中全部的excel数据，进行处理，并且放到列表中
		addr=os.path.join(self.x,'data')
		x=os.listdir(addr)
		excel_name=[]
		excel_total=[]
		for i in x:
			if ('~$' not in i) and ('.xlsx' in i):
				excel_name.append(i)
		for i in excel_name:
			x=os.path.join(addr,i)
			y=xlrd.open_workbook(x)
			z=y.sheets()[0]
			value=z.col_values(1)
			value.remove('预期结果')
			for k in range(value.count('abandon')):
				value.remove('abandon')
			for j in value:
				excel_total.append(j)
		for i in range(len(excel_total)):
			if excel_total[i] ==u'无':
				excel_total[i]=u'预期结果在页面中获得'
		return excel_total

	def write_excel(self):
		self.create_new_report()
		wb = self.create_excel()
		suite_name=self.read_xml()
		run_time=self.total_time()
		#print(run_time)
		file=open(self.log_name,'r')
		#写入第二个工作表
		ws2=wb[u'用例执行情况']
		data=file.readlines()#按行读取txt文件
		row_num=range(3,3+len(data))
		for i in range(len(data)):
			case_list=data[i].replace('\n','').split('::')
			if '' in case_list:
				case_list.remove('')
			column=['A','B','C','D','E','F','G','H','I','J']
			for j in range(len(column)):
				if 'PASS' in case_list[j]:
					ws2[column[j] + str(row_num[i])].fill=PatternFill(patternType='solid',start_color='339966')
					ws2[column[j] + str(row_num[i])] = 'PASS'
				elif 'FAIL' in case_list[j]:
					ws2[column[j] + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='FF0000')
					ws2[column[j] + str(row_num[i])] = 'FAIL'
				else:
					ws2[column[j] + str(row_num[i])] = case_list[j]
		#把得到的所有excel的预期结果重新写入到excel中
		total=self.read_all_excel()
		for i in range(len(total)):
			ws2['F'+str(i+3)]=total[i]


		#把数据写入第一个表格
		#从第8行开始写
		ws1 = wb[u'自动化测试报告']
		suite_mess=[]
		test_list=[]
		for i in suite_name:
			for j in run_time:
				if i in j:
					suite_mess.append(j)
		row_num=range(8,8+len(suite_mess))
		if len(suite_mess)==0:
			test_list=[0]
		elif len(suite_mess)>0:
			test_list=range(len(suite_mess))
		for i in test_list:
			case_list=suite_mess[i].split('::')
			ws1['A' + str(row_num[i])] = i
			ws1['B' + str(row_num[i])] = suite_name[i]
			ws1['D' + str(row_num[i])] = ws1['G' + str(row_num[i])]=case_list[1]
			ws1['E' + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='99CC00')
			ws1['E' + str(row_num[i])] = 'Y'
			ws1['F' + str(row_num[i])] = case_list[4]
			ws1['G' + str(row_num[i])] = case_list[1]
			ws1['H' + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='99CC00')
			ws1['H' + str(row_num[i])] = case_list[2]
			ws1['I' + str(row_num[i])] = case_list[3]
		#写第4行的数据
		sum_01=run_time[0].split('::')
		sum_02 = run_time[-1].split('::')
		ws1['A4'] = self.project_name
		ws1['C4'] = sum_01[1]#输入用例总数量
		ws1['D4'] = sum_01[1]#输入总执行数量
		ws1['E4'] = sum_01[2]#输入所有用例通过数量
		ws1['F4'] = sum_01[3]#输入失败用例数量
		ws1['G4'] = sum_02[0]#输入总时间
		ws1['H4'] = sum_02[1]#输入开始时间
		ws1['I4'] = sum_02[2]#输入结束时间
		wb.save(self.excel_name)#保存excel文件
		file.close()

if __name__=="__main__":
	Excel_Report(u'Project').write_excel()
