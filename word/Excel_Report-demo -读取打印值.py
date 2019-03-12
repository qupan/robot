#coding=utf-8
import xml.etree.ElementTree as ET
import xml.dom.minidom
import os,time,shutil,xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from pathlib2 import Path
try:
	import sys
	reload(sys)
	sys.setdefaultencoding('utf-8')
	print('run for python2')
except Exception as e:
	print('run for python3')


class Excel_Report():

	def __init__(self,project_name):
		self.project_name=project_name #得到项目名称
		self.report_time=time.strftime('%Y.%m.%d.%H.%M.%S') #得到文件名称
		self.x=Path().cwd().parent #得到当前目录的父目录
		self.output=self.x/'report/output.xml' #得到xml文件地址
		self.log=self.x/'report/log.html'#得到log日志文件地址
		self.report=self.x/'report/report.html'#得到report文件地址
		self.html_addr='file:///'+str(self.x/'report/report.html')#得到本地report的浏览器的地址

	def create_new_report(self):
		self.excel_addr = self.x/'report/excel_report'#得到excel报告的存放地址
		self.log_addr = self.x/'report/log'#得到log报告的存放地址
		self.sys_report_new_addr=self.x/'report/sys_report'#得到系统报告存放的路径
		self.sys_report_addr=self.x/'report/sys_report/report'#得到系统报告存放的路径
		self.sys_report_log=self.x/'report/sys_report/log'#得到系统报告存放的路径
		self.sys_report_xml=self.x/'report/sys_report/xml'#得到系统报告存放的路径
		#把所有的地址放到一个列表中，使用for循环全部创建相应的文件夹
		addr=[self.excel_addr,self.log_addr,self.sys_report_new_addr,
				self.sys_report_addr,self.sys_report_log,self.sys_report_xml]
		for i in addr:
			Path(i).mkdir(parents=True,exist_ok=True)
		x=self.sys_report_addr/(self.report_time+'-Report.html')#系统报告report文件的文件名
		y=self.sys_report_log/(self.report_time+'-log.html')#系统报告log日志文件的文件名
		z=self.sys_report_xml/(self.report_time+'-output.xml')#系统报告output文件的文件名
		# 判断当前是Window还是Linux系统，并执行相应的复制语句
		if str( type(Path()) ) == "<class 'pathlib2.WindowsPath'>":
			os.system('copy %s %s'%(self.report,x))
			os.system('copy %s %s'%(self.log,y))
			os.system('copy %s %s'%(self.output,z))
		elif str( type(Path()) ) == "<class 'pathlib2.PosixPath'>":
			os.system('cp %s %s'%(self.report,x))
			os.system('cp %s %s'%(self.log,y))
			os.system('cp %s %s'%(self.output,z))
		self.screenshot=self.x/'report/screenshot' #得到截图的路径
		self.screenshot_child=self.screenshot/self.report_time #得到以时间命名的截图子目录，的路径
		report=self.x/'report' #得到项目中report的目录地址
		Path(self.screenshot).mkdir(parents=True,exist_ok=True) #使用Pathlib2模块创建截图存放的文件夹
		Path(self.screenshot_child).mkdir(parents=True,exist_ok=True) #创建截图每天存放的文件夹
		file_name=list( report.glob('*.png') ) #使用glob函数匹配所有的截图，并返回所有名称列表
		# 使用for循环读取匹配出的png文件名称列表，移动截图到创建好的文件中
		for i in file_name:
			shutil.move(str(i),str(self.screenshot_child))
		#进行判断是哪个系统，然后把log.html日志文件复制到错误截图存放的路径下面
		if str( type(Path()) ) == "<class 'pathlib2.WindowsPath'>":
			os.system('copy %s %s'%(self.log,self.screenshot_child))
		elif str( type(Path()) ) == "<class 'pathlib2.PosixPath'>":
			os.system('cp %s %s'%(self.log,self.screenshot_child))


	def create_excel(self):
		#创建新excel，加入格式字体，填充，颜色，预置名称等

		letter=['A','C','D','E','F','G','H','I']
		self.excel_name=self.excel_addr/(self.report_time+'-Report.xlsx')#得到新的excel文件的名称
		wb=Workbook()
		ws=wb['Sheet']
		ws.title=u'自动化测试报告'
		ws=wb.create_sheet(u'用例执行情况')

		#编辑第一个工作表格信息
		ws=wb[u'自动化测试报告']
		white = Font(size=12, bold=False, name=u'等线', color="FFFFFF")#设置字体为等线属性，白色字体
		black = Font(size=12, bold=False, name=u'等线', color="000000")#设置字体为等线属性，黑色字体
		#合并单元格
		ws.merge_cells('A1:I1')
		ws.merge_cells('A2:I2')
		ws.merge_cells('A5:I5')
		ws.merge_cells('A6:D6')
		ws.merge_cells('F6:I6')
		ws.merge_cells('A4:B4')
		# 使用for循环批量合并单元格
		for i in range(7,51):
			ws.merge_cells('B%d:C%d'%(i,i))
		#更改行高和列宽
		ws.column_dimensions['A'].width=5
		ws.column_dimensions['B'].width=15
		ws.column_dimensions['C'].width=18
		ws.column_dimensions['D'].width=25
		ws.column_dimensions['E'].width=15
		ws.column_dimensions['F'].width=20
		ws.column_dimensions['G'].width=15
		ws.column_dimensions['H'].width=25
		ws.column_dimensions['I'].width=25
		for i in range(1,51):
			ws.row_dimensions[i].height=15#更改行高为15
		#添加表格线条
		thin = Side(border_style="thin", color="000000")
		border = Border(left=thin, right=thin, top=thin, bottom=thin)
		for i in ['A','B','C','D','E','F','G','H','I']:
			for j in range(1,51):
				ws[i+str(j)].border=border
				alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
				ws[i+str(j)].alignment = alignment
				ws[i+str(j)].font=black
		#添加表格颜色
		ws['A1'].fill=PatternFill(patternType='solid',start_color='333399')
		ws['A2'].fill=PatternFill(patternType='solid',start_color='99CC00')
		for i in ['A','B','C','D','E','F','G','H','I']:
			ws[i+str(3)].fill=PatternFill(patternType='solid',start_color='0066CC')
			ws[i+str(3)].font=white
		ws['C6'].fill=PatternFill(patternType='solid',start_color='0066CC')
		for i in ['A','B','D']:
			ws[i+str(7)].fill=PatternFill(patternType='solid',start_color='0066CC')
		ws['A6'].fill=PatternFill(patternType='solid',start_color='99CC00')
		ws['E6'].fill=PatternFill(patternType='solid',start_color='FF6600')
		ws['E7'].fill=PatternFill(patternType='solid',start_color='FF9900')
		ws['F6'].fill=PatternFill(patternType='solid',start_color='FFCC99')
		ws['F7'].fill=PatternFill(patternType='solid',start_color='FFCC00')
		for i in ['F','G','H','I']:
			ws[i+str(7)].fill=PatternFill(patternType='solid',start_color='0066CC')
		#写入字符
		ws['A7'].font=ws['B7'].font=ws['D7'].font=ws['A1'].font=white
		ws['A1']=u'自动化测试报告'
		ws['F6'].font=ws['E6'].font=ws['E7'].font=ws['A6'].font=ws['A2'].font=black
		ws['F7'].font=ws['G7'].font=ws['H7'].font=ws['I7'].font=white
		ws['A2']=u'自动化测试用例总计'
		ws['A6']=u'测试模块列表'
		ws['E6']=u'是否执行'
		ws['E7']=u'执行标识符'
		ws['F6']=u'执行情况'
		ws['A7']=u'NO.'
		ws['B7']=u'模块'
		ws['D7']=u'自动化测试用例数量'
		ws['F7']=u'执行时间'
		ws['G7']=u'总计执行'
		ws['H7']=u'通过数目'
		ws['I7']=u'失败数目'
		title1=[u'自动化测试项目',u'自动化测试用例数量',u'自动化测试总计执行','Pass',
				'Fail',u'测试用时',u'开始时间',u'结束时间']
		ws.merge_cells('A3:B3')
		for i in range(len(letter)):
			ws[letter[i]+str(3)]=title1[i]

		#编辑第二个表格信息
		ws1=wb[u'用例执行情况']
		#更改单元格宽度
		for i in range(1,1000):
			ws1.row_dimensions[i].height=40
		for i in ['A','B','C','D','E','F','G','H','I','J']:
			for j in range(1,1000):
				ws1[i+str(j)].border=border
				ws1[i+str(j)].font=black
				alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
				ws1[i+str(j)].alignment = alignment
		ws1.column_dimensions['A'].width=25
		ws1.column_dimensions['B'].width=40
		ws1.column_dimensions['C'].width=13
		ws1.column_dimensions['D'].width=35
		ws1.column_dimensions['E'].width=40
		ws1.column_dimensions['F'].width=25
		ws1.column_dimensions['G'].width=15
		ws1.column_dimensions['H'].width=100
		ws1.column_dimensions['I'].width=25
		ws1.column_dimensions['J'].width=25
		#合并单元格A1。添加颜色
		ws1.merge_cells('A1:J1')
		ws1['A1'].fill=PatternFill(patternType='solid',start_color='0066CC')
		ws1['A1'].alignment = alignment
		ws1['A1'].font=white
		ws1['A1']=u'自动化测试用例执行情况'
		#添加A2-G2格式和字体
		x1=['Suite Name','Case Name','Tags',u'输入数据',u'预期结果',u'实际结果','Status','Message','Start-Time','End-Time']
		x2=['A','B','C','D','E','F','G','H','I','J']
		for i in ['A','B','C','D','E','F','G','H','I','J']:
			ws1[i+str(2)].fill=PatternFill(patternType='solid',start_color='C0C0C0')
		ws1['A2']=x1[0]
		ws1['B2']=x1[1]
		ws1['C2']=x1[2]
		ws1['D2']=x1[3]
		ws1['E2']=x1[4]
		ws1['F2']=x1[5]
		ws1['G2']=x1[6]
		ws1['H2']=x1[7]
		ws1['I2']=x1[8]
		ws1['J2']=x1[9]
		return wb#返回excel对象

	def total_time(self):
		#得到运行总时间
		#设置谷歌浏览器后台运行，加入参数
		chrome_options=webdriver.ChromeOptions()
		chrome_options.add_argument('--no-sandbox')
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument('--window-size=1920,1080')
		chrome_options.add_argument('blink-settings=imagesEnabled=false')
		chrome_options.add_argument('--hide-scrollbars')
		driver=webdriver.Chrome(options=chrome_options)
		driver.get(self.html_addr)
		time.sleep(3)
		try:
			total_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(4) > td').text#得到总的时间
			start_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(2) > td').text#得到总的开始时间
			end_time=driver.find_element_by_css_selector('body > table > tbody > tr:nth-child(3) > td').text#得到总的结束时间
			tr=driver.find_elements_by_css_selector('#suite-stats > tbody>tr')
			message=[]
			s='::'
			for i in range(1,len(tr)+1):
				name=driver.find_element_by_css_selector('#suite-stats > tbody>tr:nth-child(%d)'%i)
				suit_name=name.find_element_by_css_selector('td:nth-child(1)').text#得到名字
				total_num=name.find_element_by_css_selector('td:nth-child(2)').text#得到用例集合的总数
				pass_num = name.find_element_by_css_selector('td:nth-child(3)').text#得到成功的数量
				file_num = name.find_element_by_css_selector('td:nth-child(4)').text#得到失败的数量
				run_time = name.find_element_by_css_selector('td:nth-child(5)').text#得到运行时间
				total_str=suit_name+s+total_num+s+pass_num+s+file_num+s+run_time#将所有信息使用：：符号链接在一块
				message.append(total_str)#将所有信息加入到列表中
			project =total_time+s+start_time+s+end_time#得到总的开始时间和结束时间
			message.append(project)
			driver.close()
			# 判断在哪个平台，执行关闭驱动
			if str( type(Path()) ) == "<class 'pathlib2.WindowsPath'>":
				os.system('taskkill /f /im chromedriver.exe')
			elif str( type(Path()) ) == "<class 'pathlib2.PosixPath'>":
				os.system('killall chromedriver')
		except Exception as e:
			driver.close()
			if str( type(Path()) ) == "<class 'pathlib2.WindowsPath'>":
				os.system('taskkill /f /im chromedriver.exe')
			elif str( type(Path()) ) == "<class 'pathlib2.PosixPath'>":
				os.system('killall chromedriver')
		return message

	def read_xml(self):
		dom=xml.dom.minidom.parse(str(self.output))
		root=dom.documentElement
		suite=root.getElementsByTagName("suite")#通过标签名称的到所有的用例集合
		test=root.getElementsByTagName('test')#通过标签名称的到所有的用例
		s='::'
		suit_message = []
		self.log_name=self.log_addr/(self.report_time+'-log.log')#得到新的log日志文件的地址和名称
		file=open(str(self.log_name),'w')#打开新的日志文件
		#使用for循环读取出所有的用例集
		for i in suite:
			#判断符合后缀名的文件进行使用
			if '.txt' in i.getAttribute('source') or '.robot' in i.getAttribute('source') or '.tsv' in i.getAttribute('source') or'.html' in i.getAttribute('source'):
				suite_name=i.getAttribute('name')#得到集合的名字
				suit_message.append(suite_name)
				suit_status=i.getElementsByTagName('status')#得到用例的状态
				test_case=i.getElementsByTagName('test')#得到集合下面所有的用例放入列表中
				#print(len(test_case))#打印集合下面用例的数量
				for j in test_case:
					test_name=j.getAttribute('name')#得到用例名字
					x=j.getElementsByTagName('status')#得到用例下面status标签
					case_status=x[-1].getAttribute('status')#得到用例的运行结果
					endtime=x[-1].getAttribute('endtime')#得到用例结束时间
					starttime=x[-1].getAttribute('starttime')#得到用例开始开始时间
					file.write(suite_name+s)#写入集合名字
					file.write(test_name+s)#写入集合名字
					tags=j.getElementsByTagName('tag')#得到所有tag
					if len(tags)>0:#判断是否有tags标签
						tag=[]
						for k in tags:
							tag.append(k.firstChild.data)#把所有标签加入tag空列表中
						tag=','.join(tag)#将所有的标签链接在一起
						file.write(tag+s)#写入tag名称
					elif len(tags)<=0:
						file.write(u'空'+s)

					#输入数据一栏，判断标签是否是Input Text'
					input_message=j.getElementsByTagName('kw')
					input_data=[]
					#遍历所有的kw列表
					for i in input_message:
						#判断便签名称是否第Input Text
						if i.getAttribute('name')=='Input Text':
							msg=i.getElementsByTagName('msg')#得到输入的信息
							text=msg[0].firstChild.data#操作得到数据
							file.write(text.split(' ')[2][1:-1]+',')#写入log文件里
						elif i.getAttribute('name')=='Convert Dict':
							msg=i.getElementsByTagName('msg')
							text=msg[0].firstChild.data
							file.write(text[12:38]+',')
						else:
							input_data.append(u'空')
					if len(input_data) == len(input_message):
						file.write(u'无输入数据')
					file.write(s)

					#得到实际结果和预期结果
					kw=j.getElementsByTagName('kw')
					Log=[]
					for i in kw:
						if i.getAttribute('name')=='Log':
							# 使用try捕捉异常，得到的值是空的
							try:
								#第一个是得到的信息，是实际结果，第二个是输入的信息，是预期结果
								x=i.getElementsByTagName('msg')
								y=x[0].firstChild.data#得到信息
								file.write(y+s)#写入到文件中
							except Exception as e:
								file.write(u'得到的实际值是空的'+s)
						else:
							Log.append(u'未得到结果')
					if len(Log)==(len(kw)-1):#判断是否得到的预期结果和实际结果，未得到就是脚本问题
						file.write(u'脚本运行失败，未得到实际结果' + s)

					file.write(case_status + s)#写入用例的状态

					Fail_message=j.getElementsByTagName('msg')
					for k in Fail_message:
						x=k.getAttribute('level')#如果是失败的用例，得到失败的信息
						if x=="FAIL":
							file.write(k.firstChild.data.replace('\n',',')+s)
					if case_status=='PASS':#如果是通过的用例写入无
						file.write(u'无' + s)
					file.write(starttime + s + endtime + s)#写入开始时间和结束时间
					file.write(str(len(test_case)) + s)
					file.write('\n')
		file.close()
		return suit_message

	def write_excel(self):
		#调用上面的函数，使用相应的返回值
		self.create_new_report()
		wb = self.create_excel()
		suite_name=self.read_xml()
		run_time=self.total_time()
		file=open(str(self.log_name),'r')

		#写入第二个工作表
		ws2=wb[u'用例执行情况']
		data=file.readlines()#按行读取txt文件
		row_num=range(3,3+len(data))#得到从3开始的数量，后续使用
		for i in range(len(data)):
			case_list=data[i].replace('\n','').split('::')#根据::符号进行分割所有信息，取消掉回车
			if '' in case_list:
				case_list.remove('')#移除列表中的''字符
			column=['A','B','C','D','E','F','G','H','I','J']
			#使用for循环来读取出列表的信息，并写入excel中，一共10个字段
			for j in range(len(column)):#得到列表0-9
				try:
					#使用if语句来判断每个字段是否是PASS和FAIL字符
					if 'PASS' in case_list[j]:#如果字段是pass的话写入pass字符
						ws2[column[j] + str(row_num[i])].fill=PatternFill(patternType='solid',start_color='339966')
						ws2[column[j] + str(row_num[i])] = 'PASS'
					elif 'FAIL' in case_list[j]:#如果字段是fail的话写入fail字符
						ws2[column[j] + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='FF0000')
						ws2[column[j] + str(row_num[i])] = 'FAIL'
					else:#如果字段不是pass，也不是fail字符，就写入正常的数据就可以了
						ws2[column[j] + str(row_num[i])] = case_list[j]
				except Exception as e:
					print('pass')

		#把数据写入第一个表格
		#从第8行开始写
		ws1 = wb[u'自动化测试报告']
		suite_mess=[]
		test_list=[]
		for i in suite_name:#使用for循环读出所有的集合名称
			for j in run_time:#使用for循环读出运行时间
				if i in j:#判断名称 i 是否在 j 中
					suite_mess.append(j)
		row_num=range(8,8+len(suite_mess))#得到从8开始的数字列表
		if len(suite_mess)==0:
			test_list=[0]
		elif len(suite_mess)>0:
			test_list=range(len(suite_mess))#得到以长度的列表
		for i in test_list:
			case_list=suite_mess[i].split('::')
			#写入对应的数据
			ws1['A' + str(row_num[i])] = i + 1 #写入编号
			ws1['B' + str(row_num[i])] = suite_name[i] #写入模块名称
			ws1['D' + str(row_num[i])] = ws1['G' + str(row_num[i])]=case_list[1] #写入用例集合的总数量
			ws1['E' + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='99CC00')
			ws1['E' + str(row_num[i])] = 'Y' #写入执行表示符号
			ws1['F' + str(row_num[i])] = case_list[4] #写入用例集合执行的总的时间
			ws1['G' + str(row_num[i])] = case_list[1] #写入用例集合总执行数量
			ws1['H' + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='99CC00')
			ws1['H' + str(row_num[i])] = case_list[2] #写入用例通过的数量
			ws1['I' + str(row_num[i])].fill = PatternFill(patternType='solid', start_color='FF0000')
			ws1['I' + str(row_num[i])] = case_list[3] #写入用例失败的数据
		#写第4行的数据
		sum_01=run_time[0].split('::')
		sum_02 = run_time[-1].split('::')
		ws1['A4'] = self.project_name
		ws1['C4'] = sum_01[1]#输入用例总数量
		ws1['D4'] = sum_01[1]#输入总执行数量
		ws1['E4'] = sum_01[2]#输入所有用例通过数量
		ws1['F4'] = sum_01[3]#输入失败用例数量
		ws1['G4'] = sum_02[0]#输入总时间
		ws1['H4'] = sum_02[1]#输入开始时间
		ws1['I4'] = sum_02[2]#输入结束时间
		wb.save(str(self.excel_name))#保存excel文件
		file.close()

if __name__=="__main__":
	Excel_Report(u'Project').write_excel()
